import os
import shutil
import json
import openpyxl
import PyPDF2
import time
import zipfile

token_json = json.load(open("token/token.json", "r"))


def valid_token(token):
    return token_json.get(token) is not None


def list_output(token):
    os.makedirs(f"output/{token}", exist_ok=True)
    return os.listdir(f"output/{token}")


def valid_file_name(file_name, token):
    return file_name in list_output(token)


def read_excel(token):
    test_file = openpyxl.load_workbook(f"input/{token}/test.xlsx", data_only=True)
    test_sheet = test_file["test"]
    row_count = test_sheet.max_row - 1
    if row_count > 5000:
        raise RuntimeError(f"row count limit exceed: {row_count} (> 5000)")
    return test_sheet, row_count


def export_pdf(token):
    output_path = f"output/{token}"
    if os.path.exists(output_path):
        shutil.rmtree(output_path)
    os.makedirs(output_path)

    solution = json.load(open(f"data/solution.json", "r"))
    test_sheet, _ = read_excel(token)

    name_map = {}
    test_result = []
    for row in test_sheet.iter_rows(min_row=2):
        name = row[0].value
        uid = row[1].value
        tid = row[2].value
        answer = [[f"{tid}{i + 1:02d}", row[i + 3].value] for i in range(25)]
        name_map[uid] = name
        test_result.append([uid, answer])

    for uid, test in test_result:
        merger = PyPDF2.PdfFileMerger()
        tid = test[0][0][:4]
        wrong_answer_count = 0
        for qid, answer in test:
            if solution[qid]["answer"] != answer:
                wrong_answer_count += 1
                merger.append(solution[qid]["pdf"])

        if wrong_answer_count > 0:
            timestamp = time.strftime("%Y%m%d-%H%M%S", time.localtime(time.time()))
            merge_name = f"{timestamp}-{tid}-{name_map[uid]}.pdf"
            merger.write(f"{output_path}/{merge_name}")
            merger.close()
            print(f"pdf exported: {merge_name}")
        else:
            print(f"all clear: {tid}-{name_map[uid]}")

    print("evaluation finished")


def zip_all(token):
    target_path = f"output/{token}"
    target_zip = zipfile.ZipFile(target_path, "w")
    for root, _, files in os.walk(target_path):
        for file in files:
            if file.endswith(".pdf"):
                target_zip.write(
                    os.path.join(root, file),
                    os.path.relpath(os.path.join(root, file), file),
                    compress_type=zipfile.ZIP_DEFLATED
                )
    target_zip.close()

    print("zip finished")


